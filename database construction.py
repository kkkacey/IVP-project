import googlemaps
import requests
import json
import os
import numpy as np
from openpyxl import load_workbook

gmaps = googlemaps.Client(key='AIzaSyBxAxKmbEhLrO08SmCi9M_4r6w9Y7MOER4')

basic_url =  'https://maps.googleapis.com/maps/api'
mykey = 'key=AIzaSyBxAxKmbEhLrO08SmCi9M_4r6w9Y7MOER4'

def geocoding(address):
    loc = 'address=' + address + '&'
    geocode_url = basic_url + '/geocode/json?' + loc + mykey
    
    r = requests.get(geocode_url)
    r.json()
    geocode_json = json.loads(r.text)
    
    t = geocode_json['results']
    lat = t[0]['geometry']['location']['lat']
    lng = t[0]['geometry']['location']['lng']
    coords = [lat,lng]
    return coords

#%% get camera name and translate to lat&lng

camname = list()    # store translatable camera location like '2+Ave+@+14+St,new+york,ny'
coords = np.zeros((56,2))   # store translated coordinates
os.chdir('C:\\Users\\Zhimiao\\Documents\\courses\\2-IVP\\project\\crawl\\1-10am')
for directories in os.listdir(os.getcwd()): 
    name = directories.replace(' ','+')+',new+york,ny'
    camname.append(name)
    
#    print(directories)
camname[44] = 'Broadway+@+46+St+South,new+york,ny'   # original name includes invalid character
for i in range(len(camname)):
    coords[i] = geocoding(camname[i])


#%%save camera address and coordinates
os.chdir('C:\\Users\\Zhimiao\\Documents\\courses\\2-IVP\\project\\visualization')

name2coord = load_workbook('database.xlsx')
name2coord.active = 1
sheet_translation = name2coord.active

for i in range(len(camname)):
    sheet_translation.cell(row = i + 2, column = 1).value = camname[i]
    sheet_translation.cell(row = i + 2, column = 2).value = coords[i][0]
    sheet_translation.cell(row = i + 2, column = 3).value = coords[i][1]

name2coord.save('database.xlsx')
name2coord.save('translation.xlsx')
name2coord.close()


#%% mark in map to check if coordinates are correct

#marker = 'markers=size:tiny|'
#for i in range(56):
#    latm = coords[i][0]
#    lngm = coords[i][1]
#    marker += str(latm) +','+ str(lngm) + '|'
#marker = marker[:-1]
#
#loc_para = 'center=40.753504,-73.980880&zoom=13&'
#map_para = 'size=800x400&scale=2&'
#
#map_parameters = loc_para + map_para + marker + '&'
#
#staticmap_url = basic_url + '/staticmap?' + map_parameters
#
#r = requests.get(staticmap_url[:-1])
##os.chdir('C:\\Users\\Zhimiao\\Documents\\courses\\2-IVP\\project\\visualization')
#f = open('direct translated coords' + '.png','wb')
#f.write(r.content)
#f.close()

# fail to check in the markers map!

#%% write camera location and coordinate to the database
    
os.chdir('C:\\Users\\Zhimiao\\Documents\\courses\\2-IVP\\project\\visualization')
wb_database = load_workbook('database.xlsx')
wb_database.active = 2
sheet_database = wb_database.active
#
#wb2 = load_workbook('new_database.xlsx')
#wb2.active = 1
#sheet2 = wb2.active

for i in range(len(camname)):
    sheet_database.cell(row = i + 2, column = 4).value = camname[i]
    sheet_database.cell(row = i + 2, column = 5).value = coords[i][0]
    sheet_database.cell(row = i + 2, column = 6).value = coords[i][1]

wb_database.save('database.xlsx')
wb_database.save('database - database.xlsx')
wb_database.close()

#%% 
#def search_location(camname,camadd,sheet_raw,sheet_database):
#    tcamname = str(camname).replace(' ','+')
#    index = tcamname.find('St')
#    if tcamname[:index] == str(camadd)[:index]:
#        lat = sheet2.cell(row = i + 2, column = 5).value
#        lng = sheet2.cell(row = i + 2, column = 6).value
##        lat1 = sheet2.cell(row = i + 2, column = 9).value
##        lng1 = sheet2.cell(row = i + 2, column = 10).value
##        lat2 = sheet2.cell(row = i + 2, column = 11).value
##        lng2 = sheet2.cell(row = i + 2, column = 12).value
#    return lat,lng,lat1,lng1,lat2,lng2


#%%  match raw data to camera address, and write number of cars to the database
wb_rawdata = load_workbook('database.xlsx')
wb_rawdata.active = 0
sheet_rawdata = wb_rawdata.active

wb_database = load_workbook('database.xlsx')
wb_database.active = 2
sheet_database = wb_database.active

for j in range(3):
    for i in range(43):
        name = sheet_rawdata.cell(row = 2 + j + i*3, column = 1).value
        value = sheet_rawdata.cell(row = 2 + j + i*3, column = 2).value
        sheet_rawdata.cell(row = 2 + j + i*3, column = 3).value = j
#        print(name,value)
        tname = str(name).replace(' ','+')
        index = tname.find('St')
        if index == -1:
            index = tname.find('ST')
        for k in range(56):
            camadd = sheet_database.cell(row = 2 + k, column = 4).value
            if tname[:index] == str(camadd)[:index]:
#                print(tname,camadd)
                sheet_database.cell(row = 2 + k, column = 3).value = name
                sheet_database.cell(row = 2 + k, column = 7 + j).value = value
                
#wb_rawdata.save('database.xlsx')
#wb_rawdata.save('rawdata - copy.xlsx')
#wb_rawdata.save('database.xlsx')
#wb_database.save('database - database.xlsx')
wb_database.save('database.xlsx')
wb_rawdata.close()
wb_database.close()


#%% transform address

#sheet.cell(row = 2, column = 2).value = 'hello'

#for i in range(len(road_loc)):
#    sheet.cell(row = i + 2, column = 4).value = road_loc[i]
#    index = road_loc[i].find('Ave')
#    if index == -1:
#        index = road_loc[i].find('AVE')
#    sheet.cell(row = i + 2, column = 1).value = road_loc[i][:index-1]
#wb2.save('newnew_sample_data.xlsx')
#wb2.close()
#
#
##%%
#for i in range(len(road_loc)):
#    address = sheet.cell(row = i + 2, column = 4).value
#    add = geocoding(address)
#    sheet.cell(row = i + 2, column = 5).value = add[0]
#    sheet.cell(row = i + 2, column = 6).value = add[1]
#    
#wb2.save('new_sample_data.xlsx')
#wb2.close()

#%%manually adjust the order and misleading locations in sheets



print('please verify coordinates by running drawpath.py')

#%% GeoJSON output for javascript
#wb3 = load_workbook('newnew_sample_data.xlsx')
#wb3.active = 1
#sheet3 = wb3.active
#
#paths = sheet3.iter_cols(min_col = 9, max_col = 12, min_row = 2, max_row = 57)
#for rows in paths:
#    for cell in rows:
#        print cell
##sheet3.active_cell
