import googlemaps

gmaps = googlemaps.Client(key='AIzaSyBxAxKmbEhLrO08SmCi9M_4r6w9Y7MOER4')

basic_url =  'https://maps.googleapis.com/maps/api'
mykey = 'key=AIzaSyBxAxKmbEhLrO08SmCi9M_4r6w9Y7MOER4'

#%%
from openpyxl import load_workbook

wb1 = load_workbook('newnew_sample_data.xlsx')
wb1.active = 2
sheet1 = wb1.active

wb2 = load_workbook('newnew_sample_data.xlsx')
wb2.active = 1
sheet2 = wb2.active

#%%
#path = []
#for i in range(56):
#    if sheet.cell(row = i + 2, column = 9).value==None:
#        continue
##    print(sheet.cell(row = i + 2, column = 9).value)
#    level = sheet.cell(row = i + 2, column = 8).value
#    if  level == 0:
#        color = '0x00ff00ff'
#    elif level == 1:
#        color = '0xffff00ff'
#    elif level == 2:
#        color = '0xff0000ff'
#    lat1 = sheet.cell(row = i + 2, column = 9).value
#    lng1 = sheet.cell(row = i + 2, column = 10).value
#    lat2 = sheet.cell(row = i + 2, column = 11).value
#    lng2 = sheet.cell(row = i + 2, column = 12).value
#    path.append('path=weight:3|color:' + color +'|'+ str(lat1) +','+ str(lng1) +'|'+ str(lat2) +','+ str(lng2) +'&')

#%%
def level(number):
    if  number < 70:
        dens = 0
        color = '0x00ff00ff'
    elif number > 130:
        dens = 1
        color = '0xff0000ff'
    else:
        dens = 2
        color = '0xffff00ff'
    return dens,color

#%%
#import translate

#def search_location(camloc,sheet1,sheet2):
#    lat = 0
#    lng = 0
#    lat1 = 0
#    lng2 = 0
#    lat2 = 0
#    lng2 = 0
#    for i in range(56):    
##        camloc = sheet1.cell(row = i + 2, column = 1).value
##        if camloc:
#        loc = str(camloc).replace(' ','+')
#        index = loc.find('St')
#        if loc[:index] == str(camloc)[:index]:
#            lat = sheet2.cell(row = i + 2, column = 5).value
#            lng = sheet2.cell(row = i + 2, column = 6).value
#            lat1 = sheet2.cell(row = i + 2, column = 9).value
#            lng1 = sheet2.cell(row = i + 2, column = 10).value
#            lat2 = sheet2.cell(row = i + 2, column = 11).value
#            lng2 = sheet2.cell(row = i + 2, column = 12).value
#        else:
#            continue
#        print('location not found.')
#        return lat,lng,lat1,lng1,lat2,lng2
        
def para_path(sheet1,sheet2,j):    
    path = []
    for i in range(j,56*3,3):
        camloc = sheet1.cell(row = i + 2, column = 1).value
#        print(camloc)
#        print(i)
        if camloc != None:
#            for n in range(56):    
    #        camloc = sheet1.cell(row = i + 2, column = 1).value
    #        if camloc:
            loc = str(camloc).replace('+',' ')
            index = loc.find('St')
            if index == -1:
                index = loc.find('ST')
#            print('camloc:' + camloc)
#            print(loc)
            if loc[:index] == str(camloc)[:index]:
#                print(loc)
                lat = sheet2.cell(row = i + 2, column = 5).value
                lng = sheet2.cell(row = i + 2, column = 6).value
                lat1 = sheet2.cell(row = i + 2, column = 7).value
                lng1 = sheet2.cell(row = i + 2, column = 8).value
                lat2 = sheet2.cell(row = i + 2, column = 9).value
                lng2 = sheet2.cell(row = i + 2, column = 10).value
#                lat = sheet2.cell(row = (i-j)/3 + 2, column = 5).value
#                lng = sheet2.cell(row = (i-j)/3 + 2, column = 6).value
#                lat1 = sheet2.cell(row = (i-j)/3 + 2, column = 7).value
#                lng1 = sheet2.cell(row = (i-j)/3 + 2, column = 8).value
#                lat2 = sheet2.cell(row = (i-j)/3 + 2, column = 9).value
#                lng2 = sheet2.cell(row = (i-j)/3 + 2, column = 10).value
#                matchloc = loc[:index]
#                sheet1.cell(row = i + j + 2, column = 3).value = lat
#                sheet1.cell(row = i + j + 2, column = 4).value = lng
##                sheet1.cell(row = i + 2, column = 5).value = dens
#                sheet1.cell(row = i + j + 2, column = 6).value = j
            else:
                continue
#            lat,lng,lat1,lng1,lat2,lng2 = search_location(camloc,sheet1,sheet2)
            dens,color = level(sheet1.cell(row = i + 2, column = 2).value)
            sheet1.cell(row = i + 2, column = 3).value = lat
            sheet1.cell(row = i + 2, column = 4).value = lng
            sheet1.cell(row = i + 2, column = 5).value = dens
            sheet1.cell(row = i + 2, column = 6).value = j
#            sheet1.cell(row = i + 2, column = 7).value = matchloc
            
#            print(color)
            if lat1 != None:
                path.append('path=weight:3|color:' + color +'|'+ str(lat1) +','+ str(lng1) +'|'+ str(lat2) +','+ str(lng2) +'&')
        else:
            break
    return path

#%%
marker = 'markers=size:tiny|'
for i in range(56):
    latm = sheet2.cell(row = i + 2, column = 5).value
    lngm = sheet2.cell(row = i + 2, column = 6).value
    marker += str(latm) +','+ str(lngm) + '|'
marker = marker[:-1]

#%%
loc_para = 'center=40.753504,-73.980880&zoom=13&'
map_para = 'size=800x400&scale=2&'


#%%
    
for j in range(3):
    map_parameters = loc_para + map_para + marker + '&'
    path = para_path(sheet1,sheet2,j)
    for i in range(len(path)):
        map_parameters += path[i]
    #path1 = 'path=color:0xff000080|40.7459057,-73.9739882|40.732349,-73.9849373&'
    #path2 = 'path=color:0xffff0080|40.7548422,-73.9841088|40.7592723,-73.9808828'
    #for i in range(len(add)):
    #    path += str(add[i])[2:-1] + '|'
    
    
    staticmap_url = basic_url + '/staticmap?' + map_parameters
#    print(staticmap_url[:-1])
    
    #%%
    import requests
    import os
    
    r = requests.get(staticmap_url[:-1])
    os.chdir('C:\\Users\\Zhimiao\\Documents\\courses\\2-IVP\\project\\visualization')
    f = open('map' + str(j+1) + '.png','wb')
    f.write(r.content)
    f.close()

    wb1.save('newnew_sample_data.xlsx')
wb1.close()